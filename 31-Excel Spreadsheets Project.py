import openpyxl as xl
from openpyxl.chart import BarChart, Reference


def process_workbook(filename):
    """load workbook"""
    wb = xl.load_workbook(filename)
    sheet = wb['Sheet1']

    """iterate through a rang of rows """
    for row in range(2, sheet.max_row + 1):
        """Select a specific column"""
        cell = sheet.cell(row, 3)

        """change value of the entire cell"""
        corrected_price = cell.value * 0.9

    """inserting the new values in column 4"""
    corrected_price_cell = sheet.cell(row, 4)
    corrected_price_cell.value = corrected_price

    """select a range of cells for chart"""
    values = Reference(sheet,
      min_row=2,
      max_row=sheet.max_row,  #select all columns in all rows
      min_col=4,        # select 4th column only
      max_col=4)

    """Creating bar chart object"""
    chart = BarChart()

    """passing the values of selected range to the chart object"""
    chart.add_data(values)

    """inserting the new chart to excel cell"""
    sheet.add_chart(chart, 'e2')

    """save the file"""
    wb.save(filename)


"""
# Accessing cells

cell = sheet['a1']
cell = sheet.cell(1, 1)
print(cell.value)
print(sheet.max_row)

"""